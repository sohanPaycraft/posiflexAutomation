import inspect
import logging
from datetime import datetime
import pytest
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import pandas
import json



@pytest.mark.usefixtures("setup")
class BaseClass:

    def verify_element_presence(self,ID,loc):
        wait = WebDriverWait(self.driver, 60)
        if ID=="ID":

            wait.until(EC.presence_of_element_located((AppiumBy.ID, loc)))
        else:
            wait.until(EC.presence_of_element_located((AppiumBy.CLASS_NAME, loc)))

    def create_excel(self,shiftid):
        now = datetime.now()
        now = str(now)
        now = now.replace(" ", "").replace(":", "-")
        temp = shiftid + "_"+now
        filepath = "/home/sohansagar/Documents/automate" + "/" + temp + ".xlsx"
        book = Workbook()
        sheet = book.active
        inf=["Qantity","Type","Payment","Amount","Mode"]
        sheet.append(inf)
        book.save(filepath)
        return temp

    def open_write_excel(self,temp,data):
        inf = []
        inf.append(data["Qty"])
        inf.append(data["Type"])
        inf.append(data["Mode"])
        inf.append(data["Amt"])
        inf.append(data["flow"])

        filepath ="/home/sohansagar/Documents/automate" + "/" + temp + ".xlsx"
        book = load_workbook(filepath)
        sheet =book.active
        sheet.append(inf)
        book.save(filepath)

    def load_excel(self):
        pass




    def getLogger(self):
        loggerName = inspect.stack()[1][3]
        logger = logging.getLogger(loggerName)
        fileHandler = logging.FileHandler('logfile.log')
        formatter = logging.Formatter("%(asctime)s :%(levelname)s : %(name)s :%(message)s")
        fileHandler.setFormatter(formatter)

        logger.addHandler(fileHandler)  # filehandler object

        logger.setLevel(logging.DEBUG)
        return logger
